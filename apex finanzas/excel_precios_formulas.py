"""Layout, márgenes y fórmulas del inventario Apex (única fuente de verdad)."""

from __future__ import annotations

import re

from openpyxl.utils import get_column_letter

# ── Calibración taller (base = Catálogo × 1.19, solo IVA) ───────────────────
# Pequeña: bieleta cat $18.220 → taller $30.000
_REF_BIETA_CAT = 18_220
_REF_BIETA_TALLER = 30_000
# Grande: Koleos cat $85.000 → taller $160.000
_REF_KOLEOS_CAT = 85_000
_REF_KOLEOS_TALLER = 160_000
# Público premium: bieleta cat $18.220 → público $43.900
_REF_BIETA_PUBLICO = 43_900

_CR_BIETA = round(_REF_BIETA_CAT * 1.19)
_CR_KOLEOS = round(_REF_KOLEOS_CAT * 1.19)

MARGEN_TALLER_BAJO = _CR_BIETA / _REF_BIETA_TALLER
MARGEN_TALLER_ALTO = _CR_KOLEOS / _REF_KOLEOS_TALLER
MARGEN_TALLER_MED = (MARGEN_TALLER_BAJO + MARGEN_TALLER_ALTO) / 2

MARGEN_PUBLICO_BAJO = _CR_BIETA / _REF_BIETA_PUBLICO
MARGEN_PUBLICO_ALTO = MARGEN_PUBLICO_BAJO * (MARGEN_TALLER_ALTO / MARGEN_TALLER_BAJO)
MARGEN_PUBLICO_MED = (MARGEN_PUBLICO_BAJO + MARGEN_PUBLICO_ALTO) / 2

UMBRAL_CAT_PEQUENA = 30_000
UMBRAL_CAT_GRANDE = 70_000
DESCUENTO_MAYORISTA = 0.75

# Anclas opcionales (solo si la fórmula no alcanza el estudio).
PRECIOS_MERCADO_TALLER: dict[str, int] = {}

INVENTARIO_HEADERS = [
    "REFERENCIA",
    "DESCRIPCIÓN",
    "MARCA",
    "STOCK (piezas)",
    "Catálogo / pieza",
    "Total catálogo",
    "Facturado / pieza",
    "Costo IVA / pieza",
    "Fondo moto / pieza",
    "Costo real / pieza",
    "Precio mercado taller (opc.)",
    "Precio Taller / pieza",
    "Precio Público / pieza",
    "Ganancia Taller / pieza",
    "Ganancia Público / pieza",
    "Venta total Taller",
    "Venta total Público",
    "Fondo Logística Acumulado",
    "Costo real total (stock)",
    "Ganancia total Taller",
    "Ganancia total Público",
    "Ahorro mayorista total (reinversión)",
]

COL_REF = 1
COL_DESC = 2
COL_MARCA = 3
COL_STOCK = 4
COL_CATALOGO = 5
COL_TOTAL_CATALOGO = 6
COL_FACTURADO = 7
COL_COSTO_IVA = 8
COL_FONDO_MOTO_PIEZA = 9
COL_COSTO_REAL_PIEZA = 10
COL_PRECIO_MERCADO_TALLER = 11
COL_PRECIO_TALLER = 12
COL_PRECIO_PUBLICO = 13
COL_GAN_TALLER_PIEZA = 14
COL_GAN_PUBLICO_PIEZA = 15
COL_VENTA_TOTAL_TALLER = 16
COL_VENTA_TOTAL_PUBLICO = 17
COL_FONDO_ACUMULADO = 18
COL_COSTO_REAL_TOTAL = 19
COL_GAN_TALLER_TOTAL = 20
COL_GAN_PUBLICO_TOTAL = 21
COL_AHORRO_MAYORISTA = 22

MONEY_HEADERS = frozenset(INVENTARIO_HEADERS) - {
    "REFERENCIA",
    "DESCRIPCIÓN",
    "MARCA",
    "STOCK (piezas)",
}

REF_PRODUCTO = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-]{2,}$")


def es_referencia_producto(ref: str) -> bool:
    s = ref.strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return False
    if s.startswith("←") or "NUEVO" in s.upper():
        return False
    return bool(REF_PRODUCTO.match(s))


def catalogo_efectivo(catalogo: float, facturado: float = 0) -> float:
    """Precio lista proveedor; si falta, estimar desde facturado mayorista."""
    if catalogo and catalogo > 0:
        return float(catalogo)
    if facturado and facturado > 0:
        return float(facturado) / DESCUENTO_MAYORISTA
    return 0.0


def costo_real_desde_catalogo(catalogo: float, facturado: float = 0) -> tuple[int, int, int]:
    """Base de venta = catálogo + IVA (19%). Fondo moto aparte, no entra al precio."""
    cat = catalogo_efectivo(catalogo, facturado)
    costo_iva = round(cat * 1.19)
    fondo_moto = round(costo_iva * 0.08)  # solo referencia logística, no suma al precio
    return costo_iva, fondo_moto, costo_iva


def costo_real_pieza_desde_facturado(facturado: float) -> tuple[int, int, int]:
    """Compatibilidad: costo interno pagado (facturado)."""
    costo_iva = round(float(facturado) * 1.19)
    fondo_moto = round(costo_iva * 0.08)
    costo_real = costo_iva + fondo_moto
    return costo_iva, fondo_moto, costo_real


def tramo_catalogo(catalogo: float) -> str:
    if catalogo < UMBRAL_CAT_PEQUENA:
        return "pequena"
    if catalogo <= UMBRAL_CAT_GRANDE:
        return "media"
    return "grande"


def margen_taller(catalogo: float) -> float:
    t = tramo_catalogo(catalogo)
    if t == "pequena":
        return MARGEN_TALLER_BAJO
    if t == "media":
        return MARGEN_TALLER_MED
    return MARGEN_TALLER_ALTO


def margen_publico(catalogo: float) -> float:
    t = tramo_catalogo(catalogo)
    if t == "pequena":
        return MARGEN_PUBLICO_BAJO
    if t == "media":
        return MARGEN_PUBLICO_MED
    return MARGEN_PUBLICO_ALTO


def redondear_centena(precio: float) -> int:
    return int(round(precio / 100) * 100)


def precio_taller_desde_formula(catalogo: float, facturado: float = 0) -> int:
    cat = catalogo_efectivo(catalogo, facturado)
    _, _, cr = costo_real_desde_catalogo(cat)
    return redondear_centena(cr / margen_taller(cat))


def precio_publico_desde_formula(catalogo: float, facturado: float = 0) -> int:
    cat = catalogo_efectivo(catalogo, facturado)
    _, _, cr = costo_real_desde_catalogo(cat)
    return redondear_centena(cr / margen_publico(cat))


def resolver_precio_taller(
    catalogo: float,
    facturado: float,
    precio_mercado: float | None,
) -> tuple[int, int]:
    cat = catalogo_efectivo(catalogo, facturado)
    _, _, cr = costo_real_desde_catalogo(cat)
    pt_formula = redondear_centena(cr / margen_taller(cat))
    if precio_mercado is not None and float(precio_mercado) > 0:
        pt = redondear_centena(float(precio_mercado))
    else:
        pt = pt_formula
    pt = max(pt, cr)
    return pt, pt_formula


def calcular_precios_venta(
    catalogo: float,
    facturado: float = 0,
    precio_mercado_taller: float | None = None,
) -> tuple[int, int]:
    pt, _ = resolver_precio_taller(catalogo, facturado, precio_mercado_taller)
    pp = precio_publico_desde_formula(catalogo, facturado)
    return pt, pp


def aplicar_encabezados(ws) -> None:
    for col, title in enumerate(INVENTARIO_HEADERS, 1):
        ws.cell(1, col, title)


def _col_letter(col: int) -> str:
    return get_column_letter(col)


def escribir_derivados_fila(
    ws,
    r: int,
    *,
    catalogo: float,
    facturado: float,
    stock: int,
    precio_taller: int | None,
    precio_publico: int | None,
) -> None:
    ws.cell(r, COL_TOTAL_CATALOGO, f"=D{r}*E{r}")

    if catalogo <= 0 and facturado <= 0:
        return
    if precio_taller is None or precio_publico is None:
        if catalogo:
            ws.cell(r, COL_AHORRO_MAYORISTA, f"=ROUND((E{r}-G{r})*D{r},0)")
        return

    iva, fondo, cr = costo_real_desde_catalogo(catalogo, facturado)
    pt, pp = precio_taller, precio_publico
    st = max(0, int(stock))

    ws.cell(r, COL_COSTO_IVA, iva)
    ws.cell(r, COL_FONDO_MOTO_PIEZA, fondo)
    ws.cell(r, COL_COSTO_REAL_PIEZA, cr)
    ws.cell(r, COL_GAN_TALLER_PIEZA, pt - cr)
    ws.cell(r, COL_GAN_PUBLICO_PIEZA, pp - cr)
    ws.cell(r, COL_VENTA_TOTAL_TALLER, pt * st)
    ws.cell(r, COL_VENTA_TOTAL_PUBLICO, pp * st)
    ws.cell(r, COL_FONDO_ACUMULADO, fondo * st)
    ws.cell(r, COL_COSTO_REAL_TOTAL, cr * st)
    ws.cell(r, COL_GAN_TALLER_TOTAL, pt * st - cr * st)
    ws.cell(r, COL_GAN_PUBLICO_TOTAL, pp * st - cr * st)
    ws.cell(r, COL_AHORRO_MAYORISTA, f"=ROUND((E{r}-G{r})*D{r},0)")


def aplicar_formulas_fila(ws, r: int) -> None:
    k = _col_letter(COL_PRECIO_TALLER)
    l = _col_letter(COL_PRECIO_PUBLICO)
    i = _col_letter(COL_FONDO_MOTO_PIEZA)
    j = _col_letter(COL_COSTO_REAL_PIEZA)
    o = _col_letter(COL_VENTA_TOTAL_TALLER)
    p = _col_letter(COL_VENTA_TOTAL_PUBLICO)
    rr = _col_letter(COL_COSTO_REAL_TOTAL)

    ws.cell(r, COL_TOTAL_CATALOGO, f"=D{r}*E{r}")
    ws.cell(r, COL_VENTA_TOTAL_TALLER, f"=D{r}*{k}{r}")
    ws.cell(r, COL_VENTA_TOTAL_PUBLICO, f"=D{r}*{l}{r}")
    ws.cell(r, COL_FONDO_ACUMULADO, f'=IF({i}{r}>0,{i}{r}*D{r},"")')
    ws.cell(r, COL_COSTO_REAL_TOTAL, f'=IF({j}{r}>0,{j}{r}*D{r},"")')
    ws.cell(r, COL_GAN_TALLER_TOTAL, f"=IF({rr}{r}>0,{o}{r}-{rr}{r},\"\")")
    ws.cell(r, COL_GAN_PUBLICO_TOTAL, f"=IF({rr}{r}>0,{p}{r}-{rr}{r},\"\")")
    ws.cell(r, COL_AHORRO_MAYORISTA, f"=ROUND((E{r}-G{r})*D{r},0)")
