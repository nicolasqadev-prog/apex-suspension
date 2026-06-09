"""
Aplica stock desde carga-stock-pwa.csv al inventario vivo y sincroniza la PWA.

Uso:
  1. Edita carga-stock-pwa.csv (referencia, stock_piezas)
  2. py -3 aplicar_stock_csv_a_pwa.py

Alternativa recomendada: editar STOCK en Inventario_Apex_VIVO.xlsx y ejecutar
  py -3 inventario.py sincronizar
"""

from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from inventario_config import COL_REF, COL_STOCK, INVENTARIO_VIVO, OUTPUT_STOCK_CSV, SHEET_INVENTARIO
from inventario_core import sincronizar


def main() -> None:
    if not INVENTARIO_VIVO.exists():
        raise SystemExit("No existe Inventario_Apex_VIVO.xlsx. Ejecuta: py -3 inventario.py inicializar")
    if not OUTPUT_STOCK_CSV.exists():
        raise SystemExit(f"No existe {OUTPUT_STOCK_CSV.name}. Sincroniza primero el inventario vivo.")

    stock_df = pd.read_csv(OUTPUT_STOCK_CSV)
    stock_map = {
        str(row["referencia"]).strip().upper(): max(0, int(row["stock_piezas"]))
        for _, row in stock_df.iterrows()
    }

    wb = load_workbook(INVENTARIO_VIVO)
    ws = wb[SHEET_INVENTARIO]
    actualizados = 0
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(r, COL_REF).value
        if not ref or str(ref).startswith("\u2190"):
            continue
        key = str(ref).strip().upper()
        if key in stock_map:
            ws.cell(r, COL_STOCK, stock_map[key])
            actualizados += 1

    wb.save(INVENTARIO_VIVO)
    res = sincronizar(INVENTARIO_VIVO)
    print(f"Stock aplicado desde CSV: {actualizados} referencias")
    print(f"PWA sincronizada: {res['piezas_pwa']} piezas | cambios registrados: {res['cambios_registrados']}")


if __name__ == "__main__":
    main()
