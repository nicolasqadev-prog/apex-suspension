"""
Actualiza precios en el ÚNICO inventario vivo.

Base de venta = Catálogo × 1.19 (solo IVA) + margen por tramo.

Taller por tramo (catálogo):
  pequeña < $30k  → % bieleta ($18.220 → $30.000)
  media $30k–$70k → promedio pequeña/grande
  grande > $70k   → % Koleos ($85.000 → $160.000)

Público: mismo % premium calibrado en bieleta, escalado por tramo.

Entrada y salida: Inventario_Apex_VIVO.xlsx (respaldo en backups/).
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_precios_formulas import (
    COL_CATALOGO,
    COL_COSTO_IVA,
    COL_DESC,
    COL_FACTURADO,
    COL_FONDO_ACUMULADO,
    COL_AHORRO_MAYORISTA,
    COL_COSTO_REAL_TOTAL,
    COL_GAN_PUBLICO_TOTAL,
    COL_GAN_TALLER_TOTAL,
    COL_MARCA,
    COL_PRECIO_MERCADO_TALLER,
    COL_PRECIO_PUBLICO,
    COL_PRECIO_TALLER,
    COL_REF,
    COL_STOCK,
    INVENTARIO_HEADERS,
    MARGEN_TALLER_ALTO,
    MARGEN_TALLER_BAJO,
    MARGEN_TALLER_MED,
    MONEY_HEADERS,
    PRECIOS_MERCADO_TALLER,
    aplicar_encabezados,
    aplicar_formulas_fila,
    calcular_precios_venta,
    escribir_derivados_fila,
    es_referencia_producto,
    precio_taller_desde_formula,
    redondear_centena,
    resolver_precio_taller,
)
from inventario_config import INVENTARIO_VIVO, SHEET_INVENTARIO, SHEET_RESUMEN

BASE_DIR = Path(__file__).parent
INVENTARIO_XLSX = INVENTARIO_VIVO
BACKUP_DIR = BASE_DIR / "backups"

MERCADO_FILL = PatternFill("solid", fgColor="DDEBF7")


def _col(df: pd.DataFrame, *candidatos: str) -> str | None:
    for c in candidatos:
        if c in df.columns:
            return c
    for c in df.columns:
        for cand in candidatos:
            if cand.lower() in str(c).lower():
                return c
    return None


def leer_datos_base(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=SHEET_INVENTARIO)
    col_ref = _col(df, "REFERENCIA") or df.columns[0]
    col_desc = _col(df, "DESCRIPCIÓN", "DESCRIPCION")
    col_marca = _col(df, "MARCA")
    col_stock = _col(df, "STOCK (piezas)", "STOCK")
    col_cat = _col(df, "Catálogo / pieza", "Catalogo / pieza")
    col_fact = _col(df, "Facturado / pieza")
    col_mercado = _col(df, "Precio mercado taller", "mercado taller")

    rows: list[dict] = []
    omitidas: list[str] = []
    for _, row in df.iterrows():
        ref_raw = row.get(col_ref)
        if ref_raw is None or (isinstance(ref_raw, float) and pd.isna(ref_raw)):
            continue
        ref = str(ref_raw).strip()
        if not ref or ref.lower() in ("nan", "none"):
            continue
        if not es_referencia_producto(ref):
            omitidas.append(ref)
            continue

        mercado: float | None = None
        if ref in PRECIOS_MERCADO_TALLER:
            mercado = float(PRECIOS_MERCADO_TALLER[ref])

        rows.append(
            {
                "referencia": ref,
                "descripcion": row.get(col_desc, "") if col_desc else "",
                "marca": row.get(col_marca, "") if col_marca else "",
                "stock": int(row[col_stock]) if col_stock and pd.notna(row.get(col_stock)) else 0,
                "catalogo": float(row[col_cat]) if col_cat and pd.notna(row.get(col_cat)) else 0,
                "facturado": float(row[col_fact]) if col_fact and pd.notna(row.get(col_fact)) else 0,
                "precio_mercado": mercado,
            }
        )
    out = pd.DataFrame(rows)
    out.attrs["omitidas"] = omitidas
    return out


def _estilo_celda(ws, row: int, col: int) -> None:
    title = INVENTARIO_HEADERS[col - 1]
    cell = ws.cell(row, col)
    if title in MONEY_HEADERS:
        cell.number_format = '"$"#,##0'
    elif title == "STOCK (piezas)":
        cell.number_format = "0"
        if row > 1:
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
    elif col == COL_PRECIO_MERCADO_TALLER and row > 1:
        cell.fill = MERCADO_FILL


def reescribir_hoja_resumen(wb, last_row: int) -> None:
    if SHEET_RESUMEN in wb.sheetnames:
        del wb[SHEET_RESUMEN]
    ws_r = wb.create_sheet(SHEET_RESUMEN)
    sh = SHEET_INVENTARIO
    h = get_column_letter(COL_COSTO_IVA)
    q = get_column_letter(COL_FONDO_ACUMULADO)
    r = get_column_letter(COL_COSTO_REAL_TOTAL)
    s = get_column_letter(COL_GAN_TALLER_TOTAL)
    t = get_column_letter(COL_GAN_PUBLICO_TOTAL)
    u = get_column_letter(COL_AHORRO_MAYORISTA)

    rows: list[tuple[str, str, str]] = [
        ("Total referencias activas", f'=COUNTIF(\'{sh}\'!A2:A{last_row},"?*")', "count"),
        ("Total piezas en stock", f"=SUM('{sh}'!D2:D{last_row})", "count"),
        ("", "", "sep"),
        ("── INVERSIÓN ──", "", "section"),
        (
            "Inversión facturas con IVA (stock)",
            f"=SUMPRODUCT('{sh}'!{h}2:{h}{last_row},'{sh}'!D2:D{last_row})",
            "money",
        ),
        ("Costo real total en bodega", f"=SUM('{sh}'!{r}2:{r}{last_row})", "money"),
        ("Fondo logística moto (stock)", f"=SUM('{sh}'!{q}2:{q}{last_row})", "money"),
        ("", "", "sep"),
        ("── GANANCIA SI VENDES TODO EL STOCK ──", "", "section"),
        ("Ganancia total Taller", f"=SUM('{sh}'!{s}2:{s}{last_row})", "highlight"),
        ("Ganancia total Público", f"=SUM('{sh}'!{t}2:{t}{last_row})", "highlight"),
        ("", "", "sep"),
        ("── OTROS ──", "", "section"),
        ("Ahorro mayorista (reinversión)", f"=SUM('{sh}'!{u}2:{u}{last_row})", "money"),
        ("Última actualización precios", "", "meta"),
    ]
    highlight = PatternFill("solid", fgColor="C6EFCE")
    ws_r.cell(1, 1, "Concepto").font = Font(bold=True)
    ws_r.cell(1, 2, "Valor").font = Font(bold=True)
    ws_r.column_dimensions["A"].width = 48
    ws_r.column_dimensions["B"].width = 22
    ri = 2
    for concepto, formula, kind in rows:
        ws_r.cell(ri, 1, concepto)
        if formula:
            ws_r.cell(ri, 2, formula)
        if kind in ("money", "highlight"):
            ws_r.cell(ri, 2).number_format = '"$"#,##0'
            if kind == "highlight":
                ws_r.cell(ri, 1).fill = highlight
                ws_r.cell(ri, 2).fill = highlight
                ws_r.cell(ri, 1).font = Font(bold=True)
                ws_r.cell(ri, 2).font = Font(bold=True)
        elif kind == "count":
            ws_r.cell(ri, 2).number_format = "0"
        elif kind == "section":
            ws_r.cell(ri, 1).font = Font(bold=True, color="1F4E79")
        ri += 1
    ws_r.cell(ri - 1, 2, datetime.now().strftime("%Y-%m-%d %H:%M"))


def reestructurar_workbook(destino: Path, datos: pd.DataFrame) -> dict[str, int]:
    wb = load_workbook(destino)
    ws = wb[SHEET_INVENTARIO]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    aplicar_encabezados(ws)
    for col in range(1, len(INVENTARIO_HEADERS) + 1):
        c = ws.cell(1, col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 28 if col == COL_DESC else 16

    precios = 0
    con_mercado = 0
    sin_facturado: list[str] = []
    for i, (_, row) in enumerate(datos.iterrows()):
        r = i + 2
        ws.cell(r, COL_REF, row["referencia"])
        ws.cell(r, COL_DESC, row["descripcion"])
        ws.cell(r, COL_MARCA, row["marca"])
        ws.cell(r, COL_STOCK, row["stock"])
        ws.cell(r, COL_CATALOGO, round(row["catalogo"]) if row["catalogo"] else "")

        pt: int | None = None
        pp: int | None = None
        cat = float(row["catalogo"] or 0)
        fact = float(row["facturado"] or 0)
        mercado = row.get("precio_mercado")
        if pd.notna(mercado) and float(mercado) > 0:
            ws.cell(r, COL_PRECIO_MERCADO_TALLER, redondear_centena(float(mercado)))

        if cat > 0 or fact > 0:
            if fact > 0:
                ws.cell(r, COL_FACTURADO, round(fact))
            mercado_val = float(mercado) if pd.notna(mercado) and float(mercado) > 0 else None
            pt, pp = calcular_precios_venta(cat, fact, mercado_val)
            pt_formula = precio_taller_desde_formula(cat, fact)
            if mercado_val is not None and pt != pt_formula:
                con_mercado += 1
            ws.cell(r, COL_PRECIO_TALLER, pt)
            ws.cell(r, COL_PRECIO_PUBLICO, pp)
            precios += 1
        else:
            sin_facturado.append(str(row["referencia"]))

        escribir_derivados_fila(
            ws,
            r,
            catalogo=cat,
            facturado=fact,
            stock=int(row["stock"]),
            precio_taller=pt,
            precio_publico=pp,
        )
        for col in range(1, len(INVENTARIO_HEADERS) + 1):
            _estilo_celda(ws, r, col)

    template_start = len(datos) + 2
    for j in range(30):
        r = template_start + j
        ws.cell(r, COL_STOCK, 0)
        if j == 0:
            ws.cell(r, COL_REF, "← NUEVO")
        aplicar_formulas_fila(ws, r)
        for col in range(1, len(INVENTARIO_HEADERS) + 1):
            _estilo_celda(ws, r, col)
    last_row = template_start + 29

    ultima_col = get_column_letter(len(INVENTARIO_HEADERS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ultima_col}{last_row}"
    reescribir_hoja_resumen(wb, last_row)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"
    wb.save(destino)
    wb.close()
    return {
        "filas": len(datos),
        "precios": precios,
        "con_mercado": con_mercado,
        "sin_facturado": sin_facturado,
        "last_row": last_row,
    }


def main() -> None:
    if not INVENTARIO_XLSX.exists():
        raise FileNotFoundError(f"No se encontró {INVENTARIO_XLSX}")

    datos = leer_datos_base(INVENTARIO_XLSX)
    omitidas = datos.attrs.get("omitidas", [])

    BACKUP_DIR.mkdir(exist_ok=True)
    backup = BACKUP_DIR / f"Inventario_Apex_VIVO_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    shutil.copy2(INVENTARIO_XLSX, backup)
    print(f"Respaldo: {backup.name}")

    stats = reestructurar_workbook(INVENTARIO_XLSX, datos)
    print(f"Inventario vivo: {INVENTARIO_XLSX.name}")
    print(f"Referencias: {stats['filas']} | precios: {stats['precios']}")
    print(
        f"Taller %: peq {MARGEN_TALLER_BAJO:.3f} | med {MARGEN_TALLER_MED:.3f} | "
        f"gra {MARGEN_TALLER_ALTO:.3f}  (base catalogo)"
    )
    print(f"Precio mercado (estudio) aplicado: {stats['con_mercado']} refs")
    if PRECIOS_MERCADO_TALLER:
        print(
            "Anclas mercado:",
            ", ".join(f"{k}=${v:,}" for k, v in PRECIOS_MERCADO_TALLER.items()),
        )
    if stats["sin_facturado"]:
        print(f"Sin facturado: {', '.join(stats['sin_facturado'])}")
    if omitidas:
        print(f"Filas plantilla omitidas: {len(omitidas)}")

    for ref in ("KSL-1001", "KSA-CH043", "KBJ-4008"):
        sub = datos[datos["referencia"] == ref]
        if sub.empty:
            continue
        row = sub.iloc[0]
        cat = float(row["catalogo"] or 0)
        fact = float(row["facturado"] or 0)
        mercado = row.get("precio_mercado")
        mercado_val = float(mercado) if pd.notna(mercado) and float(mercado) > 0 else None
        pt_f = precio_taller_desde_formula(cat, fact)
        pt, pp = calcular_precios_venta(cat, fact, mercado_val)
        m = f"mercado ${mercado_val:,.0f}" if mercado_val else "formula"
        print(f"{ref}: cat ${cat:,.0f} | taller ${pt:,} ({m}, calc ${pt_f:,}) | publico ${pp:,}")

    import subprocess

    subprocess.run(["py", "-3", "verificar_coherencia_excel.py"], check=False, cwd=BASE_DIR)


if __name__ == "__main__":
    main()
