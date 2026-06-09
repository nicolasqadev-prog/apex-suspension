"""
Sistema de inventario Apex — workbook vivo, movimientos, alertas, validación, sync PWA.
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from inventario_config import (
    ALERTAS_HEADERS,
    COL_CATALOGO,
    COL_DESC,
    COL_FACTURADO,
    COL_MARCA,
    COL_PRECIO_PUBLICO,
    COL_PRECIO_TALLER,
    COL_REF,
    COL_STOCK,
    EDITABLE_COLS,
    ESTADO_JSON,
    INVENTARIO_PLANTILLA,
    INVENTARIO_VIVO,
    MOVIMIENTOS_HEADERS,
    OUTPUT_PWA_JSON,
    OUTPUT_PWA_JSON_DATA,
    OUTPUT_STOCK_CSV,
    PEDIDOS_DIR,
    SHEET_ALERTAS,
    SHEET_GUIA,
    SHEET_INVENTARIO,
    SHEET_MOVIMIENTOS,
    SHEET_RESUMEN,
)

# Reutiliza lógica de precios del pedido inicial
from generar_lista_precios_pedido import (  # noqa: E402
    COSTO_OCULTO,
    DESCUENTO_TALLER_PWA_PCT,
    INVENTARIO_CSV,
    INVENTARIO_HEADERS,
    MONEY_HEADERS,
    build_dataframe,
    categoria_producto,
    export_pwa_files,
    slug_desde_referencia,
    COL_AHORRO_MAYORISTA,
    COL_FONDO_LOGISTICA,
    COL_GAN_PUBLICO_SIN_DOM,
    COL_GAN_TALLER_SIN_DOM,
)

EXTRA_ROWS = 30  # filas vacías con fórmulas para productos nuevos


def _border() -> Border:
    thin = Side(style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_row_formulas(ws, r: int) -> None:
    ws.cell(r, 6, f"=D{r}*E{r}")
    ws.cell(r, 8, f"=ROUND(D{r}*G{r}*1.19,0)")
    ws.cell(r, 11, f"=D{r}*I{r}")
    ws.cell(r, 12, f"=D{r}*J{r}")
    ws.cell(r, COL_FONDO_LOGISTICA, f"=D{r}*{COSTO_OCULTO}")
    ws.cell(r, 13, f"=H{r}+Q{r}")
    ws.cell(r, 14, f"=K{r}-M{r}")
    ws.cell(r, 15, f"=L{r}-M{r}")
    ws.cell(r, COL_AHORRO_MAYORISTA, f"=ROUND((E{r}-G{r})*D{r},0)")
    ws.cell(r, COL_GAN_TALLER_SIN_DOM, f"=K{r}-H{r}")
    ws.cell(r, COL_GAN_PUBLICO_SIN_DOM, f"=L{r}-H{r}")


def _style_inventario_cell(ws, row: int, col: int) -> None:
    title = INVENTARIO_HEADERS[col - 1]
    cell = ws.cell(row, col)
    cell.border = _border()
    if title in MONEY_HEADERS:
        cell.number_format = '"$"#,##0'
    elif title == "STOCK (piezas)":
        cell.number_format = "0"
        cell.fill = PatternFill("solid", fgColor="FFF2CC")


def _calc_precios_from_catalogo(cat: float, fact: float) -> tuple[float, float, float]:
    from generar_lista_precios_pedido import IVA, MARGEN_PUBLICO, MARGEN_TALLER

    cr = cat * (1 + IVA) + COSTO_OCULTO
    return cr / MARGEN_TALLER, cr / MARGEN_PUBLICO, cr


def _fill_inventario_sheet(ws, df: pd.DataFrame | None, data_rows: int) -> int:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_internal = PatternFill("solid", fgColor="5B5B5B")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col, title in enumerate(INVENTARIO_HEADERS, 1):
        c = ws.cell(1, col, title)
        c.fill = header_internal if col >= COL_AHORRO_MAYORISTA else header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    if df is not None:
        for idx, (_, row) in enumerate(df.iterrows()):
            r = idx + 2
            ws.cell(r, COL_REF, row["REFERENCIA"])
            ws.cell(r, COL_DESC, row["DESCRIPCIÓN DEL PRODUCTO"])
            ws.cell(r, COL_MARCA, row["MARCA DE CARRO"])
            ws.cell(r, COL_STOCK, int(row["STOCK (PIEZAS)"]))
            ws.cell(r, COL_CATALOGO, round(row["PRECIO CATÁLOGO UNIT (sin IVA)"]))
            ws.cell(r, COL_FACTURADO, round(row["PRECIO FACTURADO UNIT (sin IVA)"]))
            ws.cell(r, COL_PRECIO_TALLER, round(row["PRECIO TALLER UNIT"]))
            ws.cell(r, COL_PRECIO_PUBLICO, round(row["PRECIO PÚBLICO UNIT"]))
            _apply_row_formulas(ws, r)

    last_data = data_rows + 1
    template_start = last_data + 1
    template_end = last_data + EXTRA_ROWS

    for r in range(2, template_end + 1):
        if r > last_data:
            ws.cell(r, COL_STOCK, 0)
            _apply_row_formulas(ws, r)
        for col in range(1, len(INVENTARIO_HEADERS) + 1):
            _style_inventario_cell(ws, r, col)

    # Marca fila plantilla
    ws.cell(template_start, 1, "← NUEVO").font = Font(italic=True, color="888888")

    for col in range(COL_AHORRO_MAYORISTA, COL_GAN_PUBLICO_SIN_DOM + 1):
        ws.column_dimensions[get_column_letter(col)].outline_level = 1

    for col, title in enumerate(INVENTARIO_HEADERS, 1):
        ws.column_dimensions[get_column_letter(col)].width = 28 if title == "DESCRIPCIÓN" else 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:S{template_end}"

    # Validación STOCK >= 0
    dv_stock = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0")
    dv_stock.error = "Stock debe ser 0 o más (piezas físicas)."
    dv_stock.errorTitle = "Stock inválido"
    ws.add_data_validation(dv_stock)
    dv_stock.add(f"D2:D{template_end}")

    # Validación precios >= 0
    dv_money = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0")
    for col_letter in ("E", "G", "I", "J"):
        dv_money.add(f"{col_letter}2:{col_letter}{template_end}")
    ws.add_data_validation(dv_money)

    # Alertas visuales en stock
    ws.conditional_formatting.add(
        f"D2:D{template_end}",
        CellIsRule(operator="equal", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    ws.conditional_formatting.add(
        f"D2:D{template_end}",
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="FFEB9C")),
    )

    _protect_inventario_sheet(ws, template_end)
    return template_end


def _protect_inventario_sheet(ws, last_row: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=last_row, max_col=len(INVENTARIO_HEADERS)):
        for cell in row:
            cell.protection = Protection(locked=True)

    for r in range(2, last_row + 1):
        for c in EDITABLE_COLS:
            ws.cell(r, c).protection = Protection(locked=False)
        # Precios taller/público editables al cargar producto nuevo
        ws.cell(r, COL_PRECIO_TALLER).protection = Protection(locked=False)
        ws.cell(r, COL_PRECIO_PUBLICO).protection = Protection(locked=False)

    ws.protection.sheet = True
    ws.protection.password = "apex"
    ws.protection.enable()


def _write_resumen(wb: Workbook, last_row: int) -> None:
    if SHEET_RESUMEN in wb.sheetnames:
        del wb[SHEET_RESUMEN]
    ws_r = wb.create_sheet(SHEET_RESUMEN)
    sh = SHEET_INVENTARIO
    rows: list[tuple[str, str, str]] = [
        ("Total referencias activas", f'=COUNTIF(\'{sh}\'!A2:A{last_row},"?*")', "count"),
        ("Total piezas en stock", f"=SUM('{sh}'!D2:D{last_row})", "count"),
        ("Alertas stock bajo (≤1)", f"=COUNTIF('{sh}'!D2:D{last_row},\"<=1\")", "count"),
        ("", "", "sep"),
        ("── INVERSIÓN ──", "", "section"),
        ("Inversión en facturas (con IVA)", f"=SUM('{sh}'!H2:H{last_row})", "money"),
        ("", "", "sep"),
        ("── GANANCIA NETA REAL ──", "", "section"),
        ("★ Ganancia neta Taller", f"=SUM('{sh}'!R2:R{last_row})", "highlight"),
        ("★ Ganancia neta Público", f"=SUM('{sh}'!S2:S{last_row})", "highlight"),
        ("↳ Inversión YA descontada. NO restar otra vez.", "", "note"),
        ("", "", "sep"),
        ("── OPERATIVO ──", "", "section"),
        ("Ganancia operativa Taller", f"=SUM('{sh}'!N2:N{last_row})", "money"),
        ("Fondo logística domicilios", f"=SUM('{sh}'!Q2:Q{last_row})", "money"),
        ("Ahorro mayorista (reinversión)", f"=SUM('{sh}'!P2:P{last_row})", "money"),
        ("Última sincronización PWA", "", "meta"),
    ]
    highlight = PatternFill("solid", fgColor="C6EFCE")
    ws_r.cell(1, 1, "Concepto").font = Font(bold=True)
    ws_r.cell(1, 2, "Valor").font = Font(bold=True)
    ws_r.column_dimensions["A"].width = 48
    ws_r.column_dimensions["B"].width = 22
    ri = 2
    for concepto, formula, kind in rows:
        ws_r.cell(ri, 1, concepto)
        if formula:
            ws_r.cell(ri, 2, formula)
        if kind == "money" or kind == "highlight":
            ws_r.cell(ri, 2).number_format = '"$"#,##0'
            if kind == "highlight":
                ws_r.cell(ri, 1).fill = highlight
                ws_r.cell(ri, 2).fill = highlight
                ws_r.cell(ri, 1).font = Font(bold=True)
                ws_r.cell(ri, 2).font = Font(bold=True)
        elif kind == "count":
            ws_r.cell(ri, 2).number_format = "0"
        elif kind == "section":
            ws_r.cell(ri, 1).font = Font(bold=True, color="1F4E79")
        elif kind == "note":
            ws_r.cell(ri, 1).font = Font(italic=True, color="666666")
        ri += 1
    ws_r.cell(ri - 1, 2, datetime.now().strftime("%Y-%m-%d %H:%M"))


def _write_movimientos_sheet(wb: Workbook, movimientos: list[dict[str, Any]] | None = None) -> None:
    if SHEET_MOVIMIENTOS in wb.sheetnames:
        del wb[SHEET_MOVIMIENTOS]
    ws = wb.create_sheet(SHEET_MOVIMIENTOS)
    hfill = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(MOVIMIENTOS_HEADERS, 1):
        c = ws.cell(1, col, h)
        c.fill = hfill
        c.font = Font(color="FFFFFF", bold=True)

    movimientos = movimientos or []
    for i, m in enumerate(movimientos, 2):
        for col, key in enumerate(MOVIMIENTOS_HEADERS, 1):
            ws.cell(i, col, m.get(key, ""))

    dv = DataValidation(type="list", formula1='"Entrada,Salida,Venta,Ajuste,Compra,Sincronización"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"D2:D5000")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["H"].width = 40
    ws.freeze_panes = "A2"


def refresh_alertas_sheet(wb: Workbook, last_row: int) -> int:
    if SHEET_ALERTAS in wb.sheetnames:
        del wb[SHEET_ALERTAS]
    ws_a = wb.create_sheet(SHEET_ALERTAS)
    hfill = PatternFill("solid", fgColor="C00000")
    for col, h in enumerate(ALERTAS_HEADERS, 1):
        c = ws_a.cell(1, col, h)
        c.fill = hfill
        c.font = Font(color="FFFFFF", bold=True)

    ws_inv = wb[SHEET_INVENTARIO]
    alert_count = 0
    out_row = 2
    for r in range(2, last_row + 1):
        ref = ws_inv.cell(r, COL_REF).value
        if not ref or str(ref).startswith("←"):
            continue
        stock = ws_inv.cell(r, COL_STOCK).value
        if stock is None:
            continue
        stock = int(stock)
        if stock <= 1:
            alert_count += 1
            prio = "SIN STOCK" if stock == 0 else "STOCK BAJO"
            ws_a.cell(out_row, 1, prio)
            ws_a.cell(out_row, 2, ref)
            ws_a.cell(out_row, 3, ws_inv.cell(r, COL_DESC).value)
            ws_a.cell(out_row, 4, stock)
            ws_a.cell(
                out_row,
                5,
                "Reponer urgente" if stock == 0 else "Revisar / reponer pronto",
            )
            fill = PatternFill("solid", fgColor="FFC7CE" if stock == 0 else "FFEB9C")
            for c in range(1, 6):
                ws_a.cell(out_row, c).fill = fill
            out_row += 1

    if alert_count == 0:
        ws_a.cell(2, 1, "OK")
        ws_a.cell(2, 2, "Sin alertas")
        ws_a.cell(2, 5, "Todo el stock está por encima de 1 pieza")

    ws_a.column_dimensions["C"].width = 36
    ws_a.freeze_panes = "A2"
    return alert_count


def _write_guia_sheet(wb: Workbook) -> None:
    if SHEET_GUIA in wb.sheetnames:
        del wb[SHEET_GUIA]
    ws = wb.create_sheet(SHEET_GUIA)
    pasos = [
        ("1. Cambiar stock", "Edita columna amarilla STOCK (piezas) en hoja Inventario."),
        ("2. Registrar movimiento", "Ejecuta: py -3 registrar_movimiento.py REF -1 \"Venta taller\""),
        ("3. Sincronizar PWA", "Ejecuta: py -3 sincronizar_inventario.py"),
        ("4. Agregar producto nuevo", "En Inventario, busca fila ← NUEVO. Llena REFERENCIA, DESCRIPCIÓN, MARCA, STOCK, Catálogo y Facturado. Los precios Taller/Público se pueden calcular con el script o pegar manualmente."),
        ("5. Calcular precios auto", "py -3 inventario.py precios REFERENCIA — calcula taller/público desde catálogo."),
        ("6. Ocultar columnas grises", "Selecciona P–S → clic derecho → Ocultar."),
        ("7. Desbloquear edición", "Si Excel pide contraseña de hoja: apex"),
        ("8. Plantilla vacía", f"Usa {INVENTARIO_PLANTILLA.name} para empezar desde cero."),
        ("9. Pedido histórico", f"Los análisis de pedidos van en carpeta {PEDIDOS_DIR.name}/ — no mezclar con el VIVO."),
    ]
    ws.cell(1, 1, "Paso").font = Font(bold=True)
    ws.cell(1, 2, "Instrucción").font = Font(bold=True)
    for i, (paso, txt) in enumerate(pasos, 2):
        ws.cell(i, 1, paso)
        ws.cell(i, 2, txt)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80


def crear_workbook_vivo(df: pd.DataFrame | None = None, movimientos: list[dict] | None = None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_INVENTARIO
    data_rows = len(df) if df is not None else 0
    last_row = _fill_inventario_sheet(ws, df, data_rows if df is not None else 0)
    _write_movimientos_sheet(wb, movimientos)
    refresh_alertas_sheet(wb, last_row)
    _write_resumen(wb, last_row)
    _write_guia_sheet(wb)
    return wb


def guardar_vivo(path: Path, wb: Workbook) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def inicializar_desde_pedido(dest: Path = INVENTARIO_VIVO) -> Path:
    if dest.exists():
        raise FileExistsError(
            f"Ya existe {dest.name}. Usa --forzar para respaldar y recrear, o edita el archivo vivo."
        )
    df = build_dataframe()
    wb = crear_workbook_vivo(df, movimientos=[_mov_inicial(df)])
    guardar_vivo(dest, wb)
    export_pwa_from_workbook(wb)
    save_estado(wb)
    return dest


def _mov_inicial(df: pd.DataFrame) -> dict[str, Any]:
    return {
        "Fecha": date.today().isoformat(),
        "Hora": datetime.now().strftime("%H:%M"),
        "Referencia": "(TODAS)",
        "Tipo": "Compra",
        "Cantidad": int(df["STOCK (PIEZAS)"].sum()),
        "Stock anterior": 0,
        "Stock nuevo": int(df["STOCK (PIEZAS)"].sum()),
        "Motivo": "Carga inicial pedido 5 Junio",
        "Usuario": "sistema",
    }


def crear_plantilla(dest: Path = INVENTARIO_PLANTILLA) -> Path:
    wb = crear_workbook_vivo(df=None, movimientos=[])
    guardar_vivo(dest, wb)
    return dest


def leer_inventario_desde_wb(wb: Workbook) -> dict[str, dict[str, Any]]:
    ws = wb[SHEET_INVENTARIO]
    items: dict[str, dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(r, COL_REF).value
        if not ref or str(ref).startswith("←"):
            continue
        ref = str(ref).strip().upper()
        stock = ws.cell(r, COL_STOCK).value
        if stock is None:
            continue
        items[ref] = {
            "referencia": ref,
            "descripcion": ws.cell(r, COL_DESC).value or "",
            "marca": ws.cell(r, COL_MARCA).value or "",
            "stock": max(0, int(stock)),
            "catalogo": float(ws.cell(r, COL_CATALOGO).value or 0),
            "facturado": float(ws.cell(r, COL_FACTURADO).value or 0),
            "precio_taller": float(ws.cell(r, COL_PRECIO_TALLER).value or 0),
            "precio_publico": float(ws.cell(r, COL_PRECIO_PUBLICO).value or 0),
            "row": r,
        }
    return items


def append_movimiento(wb: Workbook, mov: dict[str, Any]) -> None:
    ws = wb[SHEET_MOVIMIENTOS]
    row = ws.max_row + 1
    if ws.cell(2, 1).value in (None, ""):
        row = 2
    for col, key in enumerate(MOVIMIENTOS_HEADERS, 1):
        ws.cell(row, col, mov.get(key, ""))


def registrar_movimiento(
    referencia: str,
    cantidad: int,
    tipo: str,
    motivo: str,
    usuario: str = "operador",
    path: Path = INVENTARIO_VIVO,
) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"No existe {path}. Ejecuta: py -3 inventario.py inicializar")

    wb = load_workbook(path)
    ws = wb[SHEET_INVENTARIO]
    items = leer_inventario_desde_wb(wb)
    ref = referencia.strip().upper()
    if ref not in items:
        raise ValueError(f"Referencia no encontrada: {ref}")

    row = items[ref]["row"]
    stock_ant = int(ws.cell(row, COL_STOCK).value or 0)
    stock_new = max(0, stock_ant + cantidad)
    ws.cell(row, COL_STOCK, stock_new)

    mov = {
        "Fecha": date.today().isoformat(),
        "Hora": datetime.now().strftime("%H:%M"),
        "Referencia": ref,
        "Tipo": tipo,
        "Cantidad": cantidad,
        "Stock anterior": stock_ant,
        "Stock nuevo": stock_new,
        "Motivo": motivo,
        "Usuario": usuario,
    }
    append_movimiento(wb, mov)

    last_row = ws.max_row
    refresh_alertas_sheet(wb, last_row)
    _write_resumen(wb, last_row)
    save_estado(wb)
    guardar_vivo(path, wb)
    return mov


def auto_calc_precios_referencia(path: Path, referencia: str) -> None:
    wb = load_workbook(path)
    items = leer_inventario_desde_wb(wb)
    ref = referencia.strip().upper()
    if ref not in items:
        raise ValueError(f"Referencia no encontrada: {ref}")
    item = items[ref]
    cat, fact = item["catalogo"], item["facturado"]
    if cat <= 0:
        raise ValueError("Catálogo / pieza debe ser mayor a 0")
    taller, publico, _ = _calc_precios_from_catalogo(cat, fact)
    ws = wb[SHEET_INVENTARIO]
    r = item["row"]
    ws.cell(r, COL_PRECIO_TALLER, round(taller))
    ws.cell(r, COL_PRECIO_PUBLICO, round(publico))
    guardar_vivo(path, wb)


def validar_filas_inventario(wb: Workbook) -> list[str]:
    errores: list[str] = []
    ws = wb[SHEET_INVENTARIO]
    refs_vistas: set[str] = set()
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(r, COL_REF).value
        if not ref or str(ref).startswith("←"):
            continue
        ref = str(ref).strip().upper()
        if ref in refs_vistas:
            errores.append(f"Fila {r}: referencia duplicada {ref}")
        refs_vistas.add(ref)
        stock = ws.cell(r, COL_STOCK).value
        if stock is not None and int(stock) < 0:
            errores.append(f"Fila {r}: stock negativo")
        cat = ws.cell(r, COL_CATALOGO).value
        if cat is not None and float(cat) < 0:
            errores.append(f"Fila {r}: catálogo negativo")
        if ref and (cat is None or float(cat or 0) == 0):
            if int(stock or 0) > 0:
                errores.append(f"Fila {r}: {ref} tiene stock pero sin precio catálogo")
    return errores


def export_pwa_from_workbook(wb: Workbook) -> int:
    import json
    from datetime import date as dt

    items = leer_inventario_desde_wb(wb)
    piezas = []
    for ref, it in items.items():
        if it["catalogo"] <= 0 and it["precio_publico"] <= 0:
            continue
        precio = int(round(it["precio_publico"])) if it["precio_publico"] > 0 else int(
            round(_calc_precios_from_catalogo(it["catalogo"], it["facturado"])[1])
        )
        nombre = str(it["descripcion"]).split(",")[0][:80] if it["descripcion"] else ref
        piezas.append(
            {
                "slug": slug_desde_referencia(ref),
                "referencia": ref,
                "nombre": nombre,
                "aplicacion": it["descripcion"],
                "categoria": categoria_producto(
                    ref, "DMB" if not ref.startswith("K") else "KTC"
                ),
                "precioLista": precio,
                "stock": it["stock"],
                "marca": it.get("marca") or "KTC",
            }
        )

    payload = {
        "meta": {
            "fuente": "inventario-apex-vivo",
            "actualizado": dt.today().isoformat(),
            "moneda": "COP",
            "descuentoTallerRecomendadoPct": DESCUENTO_TALLER_PWA_PCT,
        },
        "piezas": piezas,
    }
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    OUTPUT_PWA_JSON.write_text(text, encoding="utf-8")
    OUTPUT_PWA_JSON_DATA.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PWA_JSON_DATA.write_text(text, encoding="utf-8")

    pd.DataFrame(
        [{"referencia": p["referencia"], "stock_piezas": p["stock"]} for p in piezas]
    ).to_csv(OUTPUT_STOCK_CSV, index=False, encoding="utf-8-sig")

    return len(piezas)


def load_estado() -> dict[str, int]:
    if not ESTADO_JSON.exists():
        return {}
    return json.loads(ESTADO_JSON.read_text(encoding="utf-8"))


def save_estado(wb: Workbook) -> None:
    items = leer_inventario_desde_wb(wb)
    estado = {ref: it["stock"] for ref, it in items.items()}
    ESTADO_JSON.write_text(json.dumps(estado, ensure_ascii=False, indent=2), encoding="utf-8")


def detectar_y_registrar_cambios_stock(wb: Workbook, usuario: str = "sistema") -> int:
    prev = load_estado()
    items = leer_inventario_desde_wb(wb)
    count = 0
    for ref, it in items.items():
        old = prev.get(ref)
        new = it["stock"]
        if old is not None and old != new:
            append_movimiento(
                wb,
                {
                    "Fecha": date.today().isoformat(),
                    "Hora": datetime.now().strftime("%H:%M"),
                    "Referencia": ref,
                    "Tipo": "Sincronización",
                    "Cantidad": new - old,
                    "Stock anterior": old,
                    "Stock nuevo": new,
                    "Motivo": "Ajuste detectado al sincronizar (cambio en Excel)",
                    "Usuario": usuario,
                },
            )
            count += 1
    return count


def sincronizar(path: Path = INVENTARIO_VIVO, push_supabase: bool = True) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"No existe {path}. Ejecuta: py -3 inventario.py inicializar")

    wb = load_workbook(path)
    errores = validar_filas_inventario(wb)
    if errores:
        raise ValueError("Errores de validación:\n" + "\n".join(errores))

    cambios = detectar_y_registrar_cambios_stock(wb)
    ws = wb[SHEET_INVENTARIO]
    last_row = ws.max_row
    refresh_alertas_sheet(wb, last_row)
    _write_resumen(wb, last_row)
    n_piezas = export_pwa_from_workbook(wb)
    alertas = sum(1 for it in leer_inventario_desde_wb(wb).values() if it["stock"] <= 1)
    save_estado(wb)
    guardar_vivo(path, wb)

    result: dict[str, Any] = {
        "piezas_pwa": n_piezas,
        "cambios_registrados": cambios,
        "alertas_stock_bajo": alertas,
        "json": str(OUTPUT_PWA_JSON_DATA),
    }

    if push_supabase:
        result["supabase"] = _push_supabase()

    return result


def _push_supabase() -> str:
    import os
    import subprocess

    if not os.environ.get("SUPABASE_URL") or not os.environ.get("SUPABASE_SERVICE_ROLE_KEY"):
        return "omitido (sin variables SUPABASE en entorno)"

    repo = Path(__file__).parent.parent
    cmd = [
        "npm",
        "run",
        "sync:inventory",
        "--",
        str(OUTPUT_PWA_JSON_DATA),
    ]
    try:
        proc = subprocess.run(
            cmd,
            cwd=repo,
            capture_output=True,
            text=True,
            timeout=120,
            shell=True,
        )
        if proc.returncode != 0:
            return f"error: {proc.stderr[:200]}"
        return "ok"
    except Exception as e:
        return f"error: {e}"
