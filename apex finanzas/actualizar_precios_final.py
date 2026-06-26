"""
Actualiza precios Taller y Público en tres fuentes Excel (fórmula v2).

Costo_IVA = costo_base * 1.19
Pequeña (costo_base <= 70_000): taller = Costo_IVA/0.72, público = Costo_IVA/0.62
Grande  (costo_base >  70_000): taller = Costo_IVA/0.61, público = Costo_IVA/0.50
Redondeo a centena más cercana.

No modifica los archivos originales; escribe *_v2.xlsx en esta carpeta.
"""

from __future__ import annotations

import re
import shutil
import time
import unicodedata
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent
DESKTOP = Path.home() / "OneDrive" / "Escritorio"

UMBRAL_GRANDE = 70_000
IVA = 1.19
VACIAS_SEGUIDAS = 25
MAX_SCAN_FILAS = 50_000

ARCHIVOS = [
    {
        "entrada": BASE_DIR / "Inventario_Apex_VIVO.xlsx",
        "salida": BASE_DIR / "Inventario_Apex_VIVO_Actualizado_v2.xlsx",
        "hoja": "Inventario",
        "col_base": ("Facturado / pieza",),
        "col_taller": ("Precio Taller / pieza", "PRECIO TALLER"),
        "col_publico": ("Precio Público / pieza", "PRECIO PUBLICO", "Precio Publico / pieza"),
        "col_costo_iva": (
            "Costo IVA / pieza",
            "Total pagado factura (IVA)",
            "COSTO IVA",
        ),
    },
    {
        "entrada": DESKTOP / "YOKOMITSU.xlsx",
        "salida": BASE_DIR / "YOKOMITSU_Actualizado_v2.xlsx",
        "hoja": None,
        "col_base": ("PRECIO",),
        "col_taller": ("Precio Taller / pieza",),
        "col_publico": ("Precio Público / pieza", "Precio Publico / pieza"),
        "col_costo_iva": ("Costo IVA", "Costo_IVA"),
    },
    {
        "entrada": DESKTOP / "Lista APEX - 22 junio.xlsx",
        "salida": BASE_DIR / "Lista_APEX_Actualizada_v2.xlsx",
        "hoja": None,
        "col_base": ("Precio", "PRECIO"),
        "col_taller": ("Precio Taller / pieza",),
        "col_publico": ("Precio Público / pieza", "Precio Publico / pieza"),
        "col_costo_iva": ("Costo IVA", "Costo_IVA"),
    },
]


def _sin_acentos(s: str) -> str:
    n = unicodedata.normalize("NFD", s)
    return "".join(c for c in n if unicodedata.category(c) != "Mn")


def _norm_header(val: object) -> str:
    if val is None:
        return ""
    s = _sin_acentos(str(val).strip().upper())
    s = re.sub(r"\s+", " ", s)
    return s


def _match_header(norm: str, candidatos: tuple[str, ...]) -> bool:
    for c in candidatos:
        cn = _norm_header(c)
        if norm == cn:
            return True
    return False


def redondear_centena(precio: float) -> int:
    return int(round(precio / 100) * 100)


def calcular_precios(costo_base: float) -> tuple[int, int, int] | None:
    if costo_base is None or costo_base <= 0:
        return None
    costo_iva = costo_base * IVA
    if costo_base <= UMBRAL_GRANDE:
        pt = costo_iva / 0.72
        pp = costo_iva / 0.62
    else:
        pt = costo_iva / 0.61
        pp = costo_iva / 0.50
    return (
        int(round(costo_iva)),
        redondear_centena(pt),
        redondear_centena(pp),
    )


def _to_float(val: object) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val > 0 else None
    s = str(val).strip().replace("$", "").replace(".", "").replace(",", ".")
    if not s or s.lower() in ("nan", "none", "-"):
        return None
    try:
        f = float(s)
        return f if f > 0 else None
    except ValueError:
        return None


def _mapa_encabezados(ws, header_row: int = 1) -> dict[int, str]:
    out: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if v is not None and str(v).strip():
            out[col] = _norm_header(v)
    return out


def _buscar_columna(headers: dict[int, str], candidatos: tuple[str, ...]) -> int | None:
    for col, norm in headers.items():
        if _match_header(norm, candidatos):
            return col
    return None


def _asegurar_columna(ws, headers: dict[int, str], nombre: str, header_row: int = 1) -> int:
    col = _buscar_columna(headers, (nombre,))
    if col is not None:
        return col
    nueva = ws.max_column + 1
    cell = ws.cell(header_row, nueva, nombre)
    # Copiar estilo del encabezado vecino si existe
    if ws.max_column > 1:
        src = ws.cell(header_row, 1)
        if src.has_style:
            cell.font = copy(src.font)
            cell.fill = copy(src.fill)
            cell.alignment = copy(src.alignment)
    headers[nueva] = _norm_header(nombre)
    return nueva


def _es_fila_datos(ws, row: int, col_base: int) -> bool:
    ref = ws.cell(row, 1).value
    if ref is not None:
        rs = str(ref).strip()
        if rs and not rs.startswith("←") and "NUEVO" not in rs.upper():
            base = _to_float(ws.cell(row, col_base).value)
            if base is not None:
                return True
    base = _to_float(ws.cell(row, col_base).value)
    return base is not None


def _ultima_fila_datos(ws, col_base: int, header_row: int = 1) -> int:
    """Evita recorrer millones de filas vacias si max_row esta inflado."""
    limite = min(ws.max_row, header_row + MAX_SCAN_FILAS)
    ultima = header_row
    vacias = 0
    for row in range(header_row + 1, limite + 1):
        if _es_fila_datos(ws, row, col_base):
            ultima = row
            vacias = 0
        else:
            vacias += 1
            if vacias >= VACIAS_SEGUIDAS and row > ultima + VACIAS_SEGUIDAS:
                break
    return ultima


def procesar_hoja(ws, cfg: dict) -> dict[str, int]:
    header_row = 1
    headers = _mapa_encabezados(ws, header_row)

    col_base = _buscar_columna(headers, cfg["col_base"])
    if col_base is None:
        raise ValueError(
            f"No se encontró columna de costo base ({cfg['col_base']}) en hoja {ws.title}"
        )

    col_taller = _asegurar_columna(ws, headers, cfg["col_taller"][0], header_row)
    col_publico = _asegurar_columna(ws, headers, cfg["col_publico"][0], header_row)

    col_costo_iva = _buscar_columna(headers, cfg["col_costo_iva"])
    if col_costo_iva is None:
        col_costo_iva = _asegurar_columna(ws, headers, "Costo IVA", header_row)

    stats = {"filas": 0, "actualizadas": 0, "omitidas": 0}

    ultima = _ultima_fila_datos(ws, col_base, header_row)
    if ws.max_row > ultima + VACIAS_SEGUIDAS:
        print(
            f"    (max_row={ws.max_row}, procesando hasta fila {ultima})",
            flush=True,
        )

    for row in range(header_row + 1, ultima + 1):
        if not _es_fila_datos(ws, row, col_base):
            stats["omitidas"] += 1
            continue
        stats["filas"] += 1
        base = _to_float(ws.cell(row, col_base).value)
        if base is None:
            stats["omitidas"] += 1
            continue

        precios = calcular_precios(base)
        if precios is None:
            continue
        costo_iva, pt, pp = precios

        ws.cell(row, col_costo_iva, costo_iva)
        ws.cell(row, col_taller, pt)
        ws.cell(row, col_publico, pp)

        for c in (col_costo_iva, col_taller, col_publico):
            ws.cell(row, c).number_format = '"$"#,##0'

        stats["actualizadas"] += 1

    return stats


def procesar_archivo(cfg: dict) -> dict:
    entrada: Path = cfg["entrada"]
    salida: Path = cfg["salida"]

    if not entrada.exists():
        raise FileNotFoundError(f"No existe: {entrada}")

    t0 = time.perf_counter()
    shutil.copy2(entrada, salida)
    print(f"    copiado en {time.perf_counter() - t0:.1f}s", flush=True)

    t1 = time.perf_counter()
    wb = load_workbook(salida)
    print(f"    abierto en {time.perf_counter() - t1:.1f}s", flush=True)

    hoja_nombre = cfg["hoja"]
    if hoja_nombre:
        if hoja_nombre not in wb.sheetnames:
            raise ValueError(f"Hoja '{hoja_nombre}' no encontrada en {entrada.name}")
        hojas = [wb[hoja_nombre]]
    else:
        hojas = [wb[wb.sheetnames[0]]]

    total = {"archivo": entrada.name, "salida": salida.name, "hojas": []}
    for ws in hojas:
        st = procesar_hoja(ws, cfg)
        st["hoja"] = ws.title
        total["hojas"].append(st)

    t2 = time.perf_counter()
    wb.save(salida)
    wb.close()
    print(f"    guardado en {time.perf_counter() - t2:.1f}s", flush=True)
    return total


def main() -> None:
    print("=== actualizar_precios_final.py (fórmula v2) ===\n")
    errores: list[str] = []

    for cfg in ARCHIVOS:
        print(f">> {cfg['entrada'].name}")
        try:
            res = procesar_archivo(cfg)
            for h in res["hojas"]:
                print(
                    f"  OK {res['salida']}"
                    f" | hoja {h['hoja']}: {h['actualizadas']} precios"
                    f" ({h['filas']} filas con dato)"
                )
        except PermissionError:
            msg = f"  ERROR Archivo abierto en Excel - cierralo: {cfg['entrada']}"
            print(msg)
            errores.append(msg)
        except Exception as e:
            msg = f"  ERROR {cfg['entrada'].name}: {e}"
            print(msg)
            errores.append(msg)
        print()

    if errores:
        print("--- Errores ---")
        for e in errores:
            print(e)
        raise SystemExit(1)

    print("Listo. Archivos v2 en:", BASE_DIR)


if __name__ == "__main__":
    main()
