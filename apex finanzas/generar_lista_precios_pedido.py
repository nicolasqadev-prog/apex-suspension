"""
Genera lista de precios y stock del pedido KTC + DMB (5 Junio).
Basado en catálogo para cobrar, factura para costo real pagado.
"""

from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path

import pandas as pd
import pdfplumber
from inventario_config import PEDIDOS_DIR
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

IVA = 0.19
COSTO_OCULTO = 10_000
MARGEN_TALLER = 0.78
MARGEN_PUBLICO = 0.65
DESCUENTO_MAYORISTA_FALLBACK = 0.75

PEDIDO_XLSX = Path(r"C:\Users\Usuario\Downloads\Pedido_KTC_DMB_5Junio_completo.xlsx")
FACTURA_KTC_PDF = Path(
    r"C:\Users\Usuario\Downloads\FACTURA DE VENTA_TEXS2065504_800219028_.pdf"
)
OUTPUT_XLSX = PEDIDOS_DIR / "Pedido_5Junio_analisis.xlsx"
OUTPUT_CORREGIDO_XLSX = PEDIDOS_DIR / "Pedido_5Junio_analisis.xlsx"
OUTPUT_PWA_JSON_PEDIDO = PEDIDOS_DIR / "carga-pwa-pedido-5junio.json"
OUTPUT_PWA_JSON_DATA = (
    Path(__file__).parent.parent / "data" / "inventario-pedido-5junio.json"
)
OUTPUT_STOCK_CSV_PEDIDO = PEDIDOS_DIR / "stock-pedido-5junio.csv"
OUTPUT_PANEL_CSV_PEDIDO = PEDIDOS_DIR / "panel-pedido-5junio.csv"
INVENTARIO_CSV = Path(__file__).parent.parent / "data" / "inventario-ktc-pedido-con-precio.csv"
# La PWA aplica este % sobre precioLista para mostrar precio taller (0.65/0.78 → ~16.67 %)
DESCUENTO_TALLER_PWA_PCT = round((1 - MARGEN_PUBLICO / MARGEN_TALLER) * 100, 2)

# Factura DMB FEBP10484 — VR. TOTAL neto por línea (10% desc., sin IVA), en orden del pedido
DMB_FACTURA_NETO_ORDEN = [
    56_723.21,  # Twingo LH
    56_723.21,  # Twingo RH
    64_285.57,  # Duster/Oroch LH
    64_285.57,  # Duster/Oroch RH
    52_941.53,  # Logan/Sandero 2016 LH
    52_941.53,  # Logan/Sandero 2016 RH
    52_941.53,  # Logan II/Sandero II LH
    52_941.53,  # Logan II/Sandero II RH
    52_941.53,  # Clio II/Symbol LH
    52_941.53,  # Clio II/Symbol RH
    37_814.81,  # Aveo/Sail LH
    37_814.81,  # Aveo/Sail RH
    105_882.06,  # Captiva LH
    105_882.06,  # Captiva RH
    90_756.34,  # Mazda 3 LH
    90_756.34,  # Mazda 3 RH
]

# Líneas con OCR corrupto — precio unitario facturado confirmado (sin IVA)
KTC_MANUAL_VU = {
    "KRE-3114": 16_530,
    "KTR-4217": 23_460,
    "KTR-6015": 15_608,
    "KTR-6016": 15_608,
}


def parse_cop(value: str) -> float:
    s = re.sub(r"[^\d,.]", "", str(value).strip())
    if not s:
        raise ValueError(f"No se pudo parsear: {value}")
    if "," in s:
        parts = s.split(",")
        if len(parts[-1]) == 3:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    elif "." in s:
        parts = s.split(".")
        if len(parts[-1]) == 3:
            s = s.replace(".", "")
    return float(s)


def parse_ktc_invoice(pdf_path: Path) -> dict[str, dict]:
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            chunk = page.extract_text()
            if chunk:
                text += chunk + "\n"

    pattern = re.compile(r"(K[A-Z]{1,2}-[A-Z0-9]+)\s+(\d+)\s+(.+)")
    items: dict[str, dict] = {}

    for line in text.split("\n"):
        match = pattern.search(line)
        if not match:
            continue
        ref, cant_str, rest = match.groups()
        cant = int(cant_str)
        numbers = re.findall(r"\d{1,3}(?:[.,]\d{3})+", rest)
        if not numbers:
            continue
        total = parse_cop(numbers[-1])
        unit = total / cant
        if len(numbers) >= 2:
            unit_candidate = parse_cop(numbers[-2])
            if abs(unit_candidate * cant - total) <= 2:
                unit = unit_candidate
        items[ref] = {"cant": cant, "vu": unit, "vt": total}

    for ref, vu in KTC_MANUAL_VU.items():
        cant = items.get(ref, {}).get("cant", 2)
        items[ref] = {"cant": cant, "vu": vu, "vt": vu * cant}

    return items


def unidades_por_caja(referencia: str, proveedor: str) -> int:
    if proveedor == "DMB":
        return 1
    if referencia.startswith("KSA-"):
        return 1
    return 2


def catalogo_unitario(facturado: float, proveedor: str) -> float:
    if proveedor == "DMB":
        return facturado / 0.9
    return 0.0


def facturado_unitario(
    row: pd.Series, invoice: dict, dmb_index: int | None, catalogo: float
) -> float:
    referencia = row["REFERENCIA"]
    proveedor = row["PROVEEDOR"]

    if proveedor == "DMB":
        if dmb_index is not None and dmb_index < len(DMB_FACTURA_NETO_ORDEN):
            return DMB_FACTURA_NETO_ORDEN[dmb_index]
        val = row.get("PRECIO DMB (COP)")
        return float(val) if pd.notna(val) else 0.0

    inv = invoice.get(referencia)
    if inv:
        vu = inv["vu"]
        ratio = vu / catalogo if catalogo else 1
        if 0.72 <= ratio <= 0.78:
            return vu
        return catalogo * DESCUENTO_MAYORISTA_FALLBACK

    return catalogo * DESCUENTO_MAYORISTA_FALLBACK


def build_dataframe() -> pd.DataFrame:
    pedido = pd.read_excel(PEDIDO_XLSX, sheet_name="RESUMEN PEDIDOS", header=2)
    pedido = pedido.dropna(subset=["REFERENCIA"]).copy()
    pedido["REFERENCIA"] = pedido["REFERENCIA"].astype(str).str.strip()

    invoice = parse_ktc_invoice(FACTURA_KTC_PDF)
    rows = []
    dmb_counter = 0

    for _, item in pedido.iterrows():
        ref = item["REFERENCIA"]
        proveedor = item["PROVEEDOR"]
        upc = unidades_por_caja(ref, proveedor)

        inv = invoice.get(ref)
        if inv:
            stock = int(inv["cant"])
        else:
            stock = upc

        cajas = stock / upc
        dmb_idx = dmb_counter if proveedor == "DMB" else None
        if proveedor == "DMB":
            dmb_counter += 1

        if proveedor == "DMB":
            fact = facturado_unitario(item, invoice, dmb_idx, 0.0)
            cat = catalogo_unitario(fact, proveedor)
        else:
            cat = float(item["PRECIO TEXIM (COP)"])
            fact = facturado_unitario(item, invoice, dmb_idx, cat)

        cat_iva = cat * (1 + IVA)
        fact_iva = fact * (1 + IVA)
        ahorro_unit = cat - fact
        ahorro_iva_unit = ahorro_unit * (1 + IVA)

        cr_cobro_unit = cat_iva + COSTO_OCULTO
        cr_real_unit = fact_iva + COSTO_OCULTO

        precio_taller = cr_cobro_unit / MARGEN_TALLER
        precio_publico = cr_cobro_unit / MARGEN_PUBLICO

        gan_taller_unit = precio_taller - cr_cobro_unit
        gan_publico_unit = precio_publico - cr_cobro_unit
        gan_real_taller_unit = precio_taller - cr_real_unit
        gan_real_publico_unit = precio_publico - cr_real_unit

        rows.append(
            {
                "CAJA": item["CAJA"],
                "REFERENCIA": ref,
                "DESCRIPCIÓN DEL PRODUCTO": item["DESCRIPCIÓN DEL PRODUCTO"],
                "MARCA DE CARRO": item["MARCA DE CARRO"],
                "PROVEEDOR": proveedor,
                "CAJAS_FACTURADAS": cajas,
                "UNIDADES_POR_CAJA": upc,
                "STOCK (PIEZAS)": stock,
                "PRECIO CATÁLOGO UNIT (sin IVA)": cat,
                "PRECIO FACTURADO UNIT (sin IVA)": fact,
                "AHORRO MAYORISTA UNIT (sin IVA)": ahorro_unit,
                "PRECIO CATÁLOGO + IVA (unit)": cat_iva,
                "COSTO OCULTO UNIT": COSTO_OCULTO,
                "CR COBRO UNIT (base precios cliente)": cr_cobro_unit,
                "COSTO REAL PAGADO UNIT (factura+IVA+oculto)": cr_real_unit,
                "PRECIO TALLER UNIT": precio_taller,
                "PRECIO PÚBLICO UNIT": precio_publico,
                "GANANCIA TALLER UNIT (sobre CR cobro)": gan_taller_unit,
                "GANANCIA PÚBLICO UNIT (sobre CR cobro)": gan_publico_unit,
                "GANANCIA REAL TALLER UNIT": gan_real_taller_unit,
                "GANANCIA REAL PÚBLICO UNIT": gan_real_publico_unit,
                "TOTAL PAGADO LÍNEA (factura con IVA)": fact_iva * stock,
                "TOTAL CR COBRO LÍNEA": cr_cobro_unit * stock,
                "TOTAL COSTO REAL LÍNEA": cr_real_unit * stock,
                "TOTAL AHORRO MAYORISTA LÍNEA": ahorro_iva_unit * stock,
                "VALOR STOCK A PRECIO TALLER": precio_taller * stock,
                "VALOR STOCK A PRECIO PÚBLICO": precio_publico * stock,
            }
        )

    return pd.DataFrame(rows)


INVENTARIO_HEADERS = [
    "REFERENCIA",
    "DESCRIPCIÓN",
    "MARCA",
    "STOCK (piezas)",
    "Catálogo / pieza",
    "Total catálogo",
    "Facturado / pieza",
    "Total pagado factura (IVA)",
    "Precio Taller / pieza",
    "Precio Público / pieza",
    "Venta total Taller",
    "Venta total Público",
    "Costo real total",
    "Ganancia total Taller",
    "Ganancia total Público",
    "Ahorro mayorista total (reinversión)",
    "Fondo logística total ($10k/pza)",
    "Ganancia neta Taller (sin domicilio)",
    "Ganancia neta Público (sin domicilio)",
]

# Columnas P–S (gris): ocultar en Excel para vista limpia
COL_AHORRO_MAYORISTA = 16  # P
COL_FONDO_LOGISTICA = 17  # Q
COL_GAN_TALLER_SIN_DOM = 18  # R
COL_GAN_PUBLICO_SIN_DOM = 19  # S

MONEY_HEADERS = {
    "Catálogo / pieza",
    "Total catálogo",
    "Facturado / pieza",
    "Total pagado factura (IVA)",
    "Precio Taller / pieza",
    "Precio Público / pieza",
    "Venta total Taller",
    "Venta total Público",
    "Costo real total",
    "Ganancia total Taller",
    "Ganancia total Público",
    "Ahorro mayorista total (reinversión)",
    "Fondo logística total ($10k/pza)",
    "Ganancia neta Taller (sin domicilio)",
    "Ganancia neta Público (sin domicilio)",
}


def write_inventario_excel(df: pd.DataFrame, path: Path) -> int:
    """Excel con fórmulas: al cambiar STOCK (col D) se recalculan totales y ganancias."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Precios y Stock"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_internal_fill = PatternFill("solid", fgColor="5B5B5B")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    stock_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(INVENTARIO_HEADERS, 1):
        cell = ws.cell(1, col, title)
        cell.fill = header_internal_fill if col >= COL_AHORRO_MAYORISTA else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    for idx, (_, row) in enumerate(df.iterrows()):
        r = idx + 2
        ws.cell(r, 1, row["REFERENCIA"])
        ws.cell(r, 2, row["DESCRIPCIÓN DEL PRODUCTO"])
        ws.cell(r, 3, row["MARCA DE CARRO"])
        ws.cell(r, 4, int(row["STOCK (PIEZAS)"]))
        ws.cell(r, 5, round(row["PRECIO CATÁLOGO UNIT (sin IVA)"]))
        ws.cell(r, 7, round(row["PRECIO FACTURADO UNIT (sin IVA)"]))
        ws.cell(r, 9, round(row["PRECIO TALLER UNIT"]))
        ws.cell(r, 10, round(row["PRECIO PÚBLICO UNIT"]))

        ws.cell(r, 6, f"=D{r}*E{r}")
        ws.cell(r, 8, f"=ROUND(D{r}*G{r}*1.19,0)")
        ws.cell(r, 11, f"=D{r}*I{r}")
        ws.cell(r, 12, f"=D{r}*J{r}")
        # Costo real = lo pagado al proveedor (H) + provisión domicilios (Q)
        ws.cell(r, COL_FONDO_LOGISTICA, f"=D{r}*{COSTO_OCULTO}")
        ws.cell(r, 13, f"=H{r}+Q{r}")
        ws.cell(r, 14, f"=K{r}-M{r}")
        ws.cell(r, 15, f"=L{r}-M{r}")
        # Ahorro real por comprar más barato que catálogo (sin IVA: el impuesto no es ganancia)
        ws.cell(r, COL_AHORRO_MAYORISTA, f"=ROUND((E{r}-G{r})*D{r},0)")
        # Ganancia limpia del repuesto, sin descontar provisión domicilio
        ws.cell(r, COL_GAN_TALLER_SIN_DOM, f"=K{r}-H{r}")
        ws.cell(r, COL_GAN_PUBLICO_SIN_DOM, f"=L{r}-H{r}")

    last_row = len(df) + 1

    for row in range(2, last_row + 1):
        for col in range(1, len(INVENTARIO_HEADERS) + 1):
            cell = ws.cell(row, col)
            cell.border = border
            title = INVENTARIO_HEADERS[col - 1]
            if title in MONEY_HEADERS:
                cell.number_format = '"$"#,##0'
            elif title == "STOCK (piezas)":
                cell.number_format = "0"
                cell.fill = stock_fill

    for col, title in enumerate(INVENTARIO_HEADERS, 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 28 if title == "DESCRIPCIÓN" else 16

    for col in range(COL_AHORRO_MAYORISTA, COL_GAN_PUBLICO_SIN_DOM + 1):
        ws.column_dimensions[get_column_letter(col)].outline_level = 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:S{last_row}"

    ws_r = wb.create_sheet("Resumen")
    sh = "Lista de Precios y Stock"
    resumen_rows: list[tuple[str, str, str]] = [
        # concepto, formula/nota, tipo: count|money|note|highlight
        ("Total referencias", f"=COUNTA('{sh}'!A2:A{last_row})", "count"),
        ("Total piezas en stock", f"=SUM('{sh}'!D2:D{last_row})", "count"),
        ("", "", "sep"),
        ("── INVERSIÓN (lo que pagaste) ──", "", "section"),
        ("Inversión en facturas (con IVA)", f"=SUM('{sh}'!H2:H{last_row})", "money"),
        ("Valor catálogo total (referencia)", f"=SUM('{sh}'!F2:F{last_row})", "money"),
        ("", "", "sep"),
        ("── SI VENDES TODO EL STOCK ──", "", "section"),
        ("Recibirías vendiendo a Taller", f"=SUM('{sh}'!K2:K{last_row})", "money"),
        ("Recibirías vendiendo a Público", f"=SUM('{sh}'!L2:L{last_row})", "money"),
        ("", "", "sep"),
        ("── GANANCIA NETA REAL ──", "", "section"),
        (
            "★ Ganancia neta Taller (inversión YA descontada)",
            f"=SUM('{sh}'!R2:R{last_row})",
            "highlight",
        ),
        (
            "★ Ganancia neta Público (inversión YA descontada)",
            f"=SUM('{sh}'!S2:S{last_row})",
            "highlight",
        ),
        (
            "↳ Es Venta total − Factura pagada. NO le restes la inversión otra vez.",
            "",
            "note",
        ),
        ("", "", "sep"),
        ("── DESPUÉS DE SEPARAR DOMICILIOS ──", "", "section"),
        ("Provisión domicilios ($10k × piezas)", f"=SUM('{sh}'!Q2:Q{last_row})", "money"),
        ("Ganancia operativa Taller (post-domicilio)", f"=SUM('{sh}'!N2:N{last_row})", "money"),
        ("Ganancia operativa Público (post-domicilio)", f"=SUM('{sh}'!O2:O{last_row})", "money"),
        ("", "", "sep"),
        ("── FONDOS APARTE (reinversión) ──", "", "section"),
        ("Ahorro mayorista (compraste más barato)", f"=SUM('{sh}'!P2:P{last_row})", "money"),
        ("Costo real total (factura + domicilio)", f"=SUM('{sh}'!M2:M{last_row})", "money"),
    ]
    highlight_fill = PatternFill("solid", fgColor="C6EFCE")
    section_font = Font(bold=True, color="1F4E79")

    ws_r.cell(1, 1, "Concepto").font = Font(bold=True)
    ws_r.cell(1, 2, "Valor").font = Font(bold=True)
    ws_r.column_dimensions["A"].width = 48
    ws_r.column_dimensions["B"].width = 18

    row_idx = 2
    for concepto, formula, kind in resumen_rows:
        ws_r.cell(row_idx, 1, concepto)
        if formula:
            ws_r.cell(row_idx, 2, formula)
        if kind == "count":
            ws_r.cell(row_idx, 2).number_format = "0"
        elif kind == "money":
            ws_r.cell(row_idx, 2).number_format = '"$"#,##0'
        elif kind == "highlight":
            ws_r.cell(row_idx, 1).font = Font(bold=True)
            ws_r.cell(row_idx, 2).font = Font(bold=True)
            ws_r.cell(row_idx, 1).fill = highlight_fill
            ws_r.cell(row_idx, 2).fill = highlight_fill
            ws_r.cell(row_idx, 2).number_format = '"$"#,##0'
        elif kind == "section":
            ws_r.cell(row_idx, 1).font = section_font
        elif kind == "note":
            ws_r.cell(row_idx, 1).font = Font(italic=True, color="666666")
        row_idx += 1

    ws_n = wb.create_sheet("Notas")
    notas = [
        ("STOCK (piezas)", "Columna amarilla. Cámbiala al hacer inventario (3, 5, 0…). Todo se recalcula solo."),
        ("Catálogo / pieza", "Precio de lista del proveedor. Base para cobrar (NO incluye descuento mayorista)."),
        ("Total catálogo", "Fórmula: STOCK × Catálogo / pieza"),
        ("Facturado / pieza", "Lo que pagaste realmente (mayorista). Solo para tu control interno."),
        ("Precios Taller / Público", "Calculados con catálogo + IVA + $10.000 oculto. Margen 22% y 35%."),
        (
            "Costo real total (col M)",
            "Fórmula: H + Q. Lo pagado al proveedor con IVA + provisión domicilio ($10k × stock).",
        ),
        (
            "Ganancia Taller/Público (col N/O)",
            "Venta total − Costo real (M). Ya separa los $10k/pza para logística.",
        ),
        (
            "Columnas P–S (gris)",
            "Ocultar con clic derecho > Ocultar. Análisis interno y reinversión.",
        ),
        (
            "Ahorro mayorista (col P)",
            "Fórmula: (Catálogo − Facturado) × STOCK, sin IVA.",
        ),
        (
            "Fondo logística (col Q)",
            "Fórmula: STOCK × $10.000. Provisión domicilios; va aparte de la ganancia del repuesto.",
        ),
        (
            "Ganancia neta sin domicilio (col R/S)",
            "Fórmula: Venta total − Total pagado factura (H). La inversión YA está descontada. "
            "NO le restes la inversión otra vez en el Resumen.",
        ),
        (
            "Resumen (hoja verde)",
            "Las filas ★ muestran ganancia neta real. Comparar con inversión da un error "
            "(parecería que solo ganas ~$225k cuando en realidad son millones).",
        ),
        ("Mercancía nueva", "Agrega una fila copiando fórmulas de otra referencia y cambia datos unitarios."),
        ("PWA", f"Subir carga-pwa-catalogo.json. Descuento taller recomendado: {DESCUENTO_TALLER_PWA_PCT}%."),
    ]
    ws_n.cell(1, 1, "Regla").font = Font(bold=True)
    ws_n.cell(1, 2, "Descripción").font = Font(bold=True)
    for idx, (regla, desc) in enumerate(notas, 2):
        ws_n.cell(idx, 1, regla)
        ws_n.cell(idx, 2, desc)
    ws_n.column_dimensions["A"].width = 22
    ws_n.column_dimensions["B"].width = 70

    wb.save(path)
    return last_row


def style_workbook(path: Path) -> None:
    """Compatibilidad: el inventario vivo ya se guarda con estilos en write_inventario_excel."""
    pass


def categoria_producto(referencia: str, proveedor: str) -> str:
    if referencia.startswith("KSA-"):
        return "Amortiguadores"
    if proveedor == "DMB":
        return "Tijeras y Brazos"
    if referencia.startswith("KSL-"):
        return "Bujes y Bieletas"
    if referencia.startswith(("KTR-", "KRE-", "KBJ-")):
        return "Terminales y Rótulas"
    return "Suspensión"


def slug_desde_referencia(referencia: str) -> str:
    return referencia.lower().strip().replace(" ", "-")


def cargar_nombres_csv() -> dict[str, dict[str, str]]:
    if not INVENTARIO_CSV.exists():
        return {}
    csv_df = pd.read_csv(INVENTARIO_CSV)
    out: dict[str, dict[str, str]] = {}
    for _, row in csv_df.iterrows():
        ref = str(row["referencia"]).strip()
        out[ref] = {
            "nombre": str(row.get("nombre", "")).strip(),
            "aplicacion": str(row.get("descripcion", "")).strip(),
        }
    return out


def export_pwa_files(df: pd.DataFrame) -> None:
    nombres = cargar_nombres_csv()
    piezas = []

    for _, row in df.iterrows():
        ref = row["REFERENCIA"]
        meta = nombres.get(ref, {})
        nombre = meta.get("nombre") or str(row["DESCRIPCIÓN DEL PRODUCTO"]).split(",")[0][:80]
        aplicacion = meta.get("aplicacion") or str(row["DESCRIPCIÓN DEL PRODUCTO"])
        proveedor = row["PROVEEDOR"]
        marca = "DMB" if proveedor == "DMB" else "KTC"

        piezas.append(
            {
                "slug": slug_desde_referencia(ref),
                "referencia": ref,
                "nombre": nombre,
                "aplicacion": aplicacion,
                "categoria": categoria_producto(ref, proveedor),
                "precioLista": int(round(row["PRECIO PÚBLICO UNIT"])),
                "stock": int(row["STOCK (PIEZAS)"]),
                "marca": marca,
            }
        )

    payload = {
        "meta": {
            "fuente": "pedido-ktc-dmb-5-junio",
            "actualizado": date.today().isoformat(),
            "moneda": "COP",
            "nota": (
                "precioLista = precio PÚBLICO (catálogo + IVA + costo oculto, margen 35%). "
                "El descuento mayorista del proveedor NO se traslada al cliente. "
                f"En la PWA, el precio taller se obtiene con {DESCUENTO_TALLER_PWA_PCT}% de descuento "
                "sobre precioLista (equivalente a margen 22% sobre CR de cobro)."
            ),
            "descuentoTallerRecomendadoPct": DESCUENTO_TALLER_PWA_PCT,
        },
        "piezas": piezas,
    }

    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    PEDIDOS_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_PWA_JSON_PEDIDO.write_text(json_text, encoding="utf-8")
    OUTPUT_PWA_JSON_DATA.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PWA_JSON_DATA.write_text(json_text, encoding="utf-8")

    panel = pd.DataFrame(
        {
            "referencia": df["REFERENCIA"],
            "descripcion": df["DESCRIPCIÓN DEL PRODUCTO"],
            "stock_piezas": df["STOCK (PIEZAS)"],
            "precio_publico_pwa": df["PRECIO PÚBLICO UNIT"].round(0).astype(int),
            "precio_taller_referencia": df["PRECIO TALLER UNIT"].round(0).astype(int),
            "marca": df["MARCA DE CARRO"],
            "activo": True,
        }
    )
    panel.to_csv(OUTPUT_PANEL_CSV_PEDIDO, index=False, encoding="utf-8-sig")

    df[["REFERENCIA", "STOCK (PIEZAS)"]].rename(
        columns={"REFERENCIA": "referencia", "STOCK (PIEZAS)": "stock_piezas"}
    ).to_csv(OUTPUT_STOCK_CSV_PEDIDO, index=False, encoding="utf-8-sig")


def build_resumen(df: pd.DataFrame) -> pd.DataFrame:
    stock_piezas = int(df["STOCK (PIEZAS)"].sum())
    return pd.DataFrame(
        [
            ("Total referencias", len(df)),
            ("Total piezas en stock", stock_piezas),
            ("Total pagado facturas (con IVA)", df["TOTAL PAGADO LÍNEA (factura con IVA)"].sum()),
            ("Total CR cobro (catálogo+IVA+oculto)", df["TOTAL CR COBRO LÍNEA"].sum()),
            ("Total costo real (factura+IVA+oculto)", df["TOTAL COSTO REAL LÍNEA"].sum()),
            ("Total ahorro mayorista (no trasladado al cliente)", df["TOTAL AHORRO MAYORISTA LÍNEA"].sum()),
            ("Valor stock a precio taller", df["VALOR STOCK A PRECIO TALLER"].sum()),
            ("Valor stock a precio público", df["VALOR STOCK A PRECIO PÚBLICO"].sum()),
            (
                "Ganancia potencial real si vende todo a taller",
                (
                    (df["PRECIO TALLER UNIT"] - df["COSTO REAL PAGADO UNIT (factura+IVA+oculto)"])
                    * df["STOCK (PIEZAS)"]
                ).sum(),
            ),
            (
                "Ganancia potencial real si vende todo a público",
                (
                    (df["PRECIO PÚBLICO UNIT"] - df["COSTO REAL PAGADO UNIT (factura+IVA+oculto)"])
                    * df["STOCK (PIEZAS)"]
                ).sum(),
            ),
        ],
        columns=["Concepto", "Valor"],
    )


def main() -> None:
    PEDIDOS_DIR.mkdir(parents=True, exist_ok=True)
    df = build_dataframe()

    export_pwa_files(df)

    for path in (OUTPUT_XLSX,):
        try:
            write_inventario_excel(df, path)
            print(f"Excel inventario: {path}  ({len(INVENTARIO_HEADERS)} columnas con formulas)")
        except PermissionError:
            print(f"Aviso: cierra {path.name} en Excel para regenerarlo.")

    sample = df[df["REFERENCIA"] == "KTR-4015"].iloc[0]
    print(f"PWA JSON pedido: {OUTPUT_PWA_JSON_PEDIDO}")
    print(f"PWA JSON data:   {OUTPUT_PWA_JSON_DATA}")
    print(f"Stock CSV:       {OUTPUT_STOCK_CSV_PEDIDO}")
    print(f"Panel CSV:       {OUTPUT_PANEL_CSV_PEDIDO}")
    print(f"Referencias: {len(df)} | Piezas en stock: {int(df['STOCK (PIEZAS)'].sum())}")
    print("Inventario vivo: py -3 inventario.py inicializar")
    print(
        "KTR-4015 - stock:",
        int(sample["STOCK (PIEZAS)"]),
        "| taller:",
        round(sample["PRECIO TALLER UNIT"]),
        "| publico:",
        round(sample["PRECIO PÚBLICO UNIT"]),
        "| precioLista PWA:",
        int(round(sample["PRECIO PÚBLICO UNIT"])),
    )


if __name__ == "__main__":
    main()
